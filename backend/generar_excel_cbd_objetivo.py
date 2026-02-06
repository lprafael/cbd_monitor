#!/usr/bin/env python3
"""
Script para generar Excel con CBD_objetivo por empresa, fecha y hora.

CBD_objetivo = cantidad de buses que cada empresa debe sacar para alcanzar 
IFO 100% (Nivel A) en cada hora/franja.

Fórmula: IFO_hora = CBD_objetivo / CBD_observado * 100
Para IFO 100%: CBD_objetivo = ceil(1.00 * B_dist_ajustado)

Uso:
    python generar_excel_cbd_objetivo.py [--ano 2026] [--salida cbd_objetivo.xlsx]
    python generar_excel_cbd_objetivo.py --verificacion   # Envía correo al director con todas las empresas
    python generar_excel_cbd_objetivo.py --notificacion   # Envía correo a cada empresa con su CBD objetivo

Requisitos: pip install psycopg2-binary openpyxl python-dotenv

Configuración .env:
    DB_*, EMAIL_HOST, EMAIL_PORT, EMAIL_USER, EMAIL_PASSWORD, EMAIL_FROM,
    EMAIL_DIRECTOR (para --verificacion)
"""

import sys
import os
import argparse
import math
import smtplib
from datetime import date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Cargar .env (buscar en backend/ y en raíz del proyecto)
try:
    from dotenv import load_dotenv  # type: ignore[import-untyped]
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    for _env_path in [
        os.path.join(_script_dir, ".env"),
        os.path.join(_script_dir, "..", ".env"),
    ]:
        if os.path.exists(_env_path):
            load_dotenv(_env_path)
            break
except ImportError:
    pass

import psycopg2  # type: ignore[import-untyped]
from psycopg2.extras import RealDictCursor  # type: ignore[import-untyped]


# =============================================================================
# CONFIGURACIÓN Y UTILIDADES
# =============================================================================

def get_db_cursor():
    """Obtiene cursor de conexión a la BD usando variables de entorno."""
    conn = psycopg2.connect(
        host=os.getenv("DB_HOST", "localhost"),
        port=os.getenv("DB_PORT", "5432"),
        database=os.getenv("DB_NAME", ""),
        user=os.getenv("DB_USER", ""),
        password=os.getenv("DB_PASSWORD", ""),
    )
    return conn.cursor(cursor_factory=RealDictCursor), conn


# =============================================================================
# TEXTO PARA EMPRESAS - Plantilla explicativa del CBD Objetivo
# =============================================================================

TEXTO_BASE_EMPRESA = """
Estimada empresa {eot_nombre},

La Dirección de Metropolitana de Transporte (DMT) les comunica el CBD Objetivo de su empresa
para los días comprendidos entre el domingo 1 y el jueves 5 de febrero de {ano}, a efectos de
obtener el Nivel A del Índice de Flota Operativa (IFO 100%) establecido en la normativa vigente.

¿QUÉ ES EL CBD OBJETIVO?
El CBD (Cantidad de Buses Diferentes) Objetivo representa la cantidad de buses distintos que su
empresa puede operar en cada hora/franja para cumplir con el Nivel A del IFO 100%.

FECHAS DE ANÁLISIS (CBD Histórico):
El cálculo se realizó utilizando como referencia el promedio de CBD observado en los mismos días de la
semana de las últimas 4 semanas anteriores. Para cada fecha objetivo, se tomaron las siguientes fechas
históricas equivalentes:

{fechas_historicas_texto}

Este método asegura que la referencia sea comparable (lunes con lunes, martes con martes, etc.), tal
como lo establece la Resolución GVMT N° 120/2025.

CBD TOPEADO AL MÍNIMO NORMATIVO:
El CBD Objetivo calculado no puede ser inferior al CBD Mínimo establecido en la tabla normativa
(control_metricas.cbd_parametros_minimos) para cada franja horaria. Esto garantiza que
nunca se solicite a una empresa una operativa por debajo de lo exigido por la norma.

MARCO NORMATIVO:
El presente análisis se basa estrictamente en lo dispuesto por la Resolución GVMT N° 120/2025,
específicamente en:
• Art. 3.1 y Anexo 1: Índice de Flota Operativa (IFO)
• Art. 5: Niveles de servicio (Nivel A: 90%-100%, Nivel B: 80%-89%, Nivel C: <80%)
• Anexo 1.1: Fórmula IFO por hora = (CBD observado / B_dist ajustado) × 100

OPERATIVA A CUMPLIR:
A continuación se detalla el CBD Objetivo por fecha y hora. Cada valor indica la cantidad de buses
diferentes que su empresa debe tener operando en esa hora para cumplir el Nivel A al 100%.
Se deben considerar las franjas operativas según el tipo de día (Laboral, Sábado, Domingo/Feriado).
Independientemente del CBD por hora, también se debe cumplir el CBD Franja (cantidad mínima de buses
distintos en el conjunto de horas de cada franja), establecido en la Resolución GVMT N° 120/2025.

{cbd_tabla_html}

Por cualquier consulta, comunicarse con la Dirección de Metropolitana de Transporte.

Saludos cordiales,
DMT - Centro de Control y Monitoreo
"""

TEXTO_VERIFICACION_DIRECTOR = """
Se adjunta a continuación el CBD Objetivo de todas las empresas permisionarias para los días 
1 al 5 de febrero de {ano}. Este informe permite la verificación del cumplimiento operativo previsto.

Fechas históricas utilizadas para el análisis (por cada fecha objetivo):
{fechas_historicas_texto}

Metodología: Resolución GVMT N° 120/2025. CBD Objetivo = mínimo para alcanzar IFO 100% (Nivel A).
El CBD Objetivo está topeado al CBD Mínimo normativo y no puede ser inferior.

{cbd_tabla_todas_html}
"""


# =============================================================================
# CÁLCULOS CBD
# =============================================================================

def _es_fecha_atipica(cursor, fecha: date) -> bool:
    cursor.execute(
        "SELECT 1 FROM control_metricas.dias_atipicos WHERE fecha = %s LIMIT 1",
        (fecha,)
    )
    return cursor.fetchone() is not None


def _ajustar_fecha_no_atipica(cursor, fecha: date, fechas_usadas: set) -> date:
    fecha_ajustada = fecha
    while _es_fecha_atipica(cursor, fecha_ajustada) or fecha_ajustada in fechas_usadas:
        fecha_ajustada -= timedelta(weeks=1)
    return fecha_ajustada


def _get_fechas_base_referencia(fecha: date) -> list:
    month, year, weekday = fecha.month, fecha.year, fecha.weekday()
    ref_month, ref_year = None, year
    if month == 1: ref_month, ref_year = 11, year - 1
    elif month == 12: ref_month = 11
    elif month == 3: ref_month, ref_year = 11, year - 1
    if ref_month:
        fechas = []
        d = date(ref_year, ref_month, 1)
        while d.weekday() != weekday:
            d += timedelta(days=1)
        while len(fechas) < 4 and d.month == ref_month:
            fechas.append(d)
            d += timedelta(weeks=1)
        if len(fechas) >= 4:
            return fechas[:4]
    return [fecha - timedelta(weeks=i) for i in range(1, 5)]


def get_fechas_referencia(cursor, fecha: date) -> list:
    fechas_base = _get_fechas_base_referencia(fecha)
    fechas_ajustadas = []
    fechas_usadas = set()
    for fecha_ref in fechas_base:
        fecha_ajustada = _ajustar_fecha_no_atipica(cursor, fecha_ref, fechas_usadas)
        fechas_ajustadas.append(fecha_ajustada)
        fechas_usadas.add(fecha_ajustada)
    return fechas_ajustadas


def get_factores_ajuste_acumulados(cursor, fecha: date) -> tuple:
    factor_total, ajustes = 1.0, []
    month, dia_semana = fecha.month, fecha.weekday()
    if month == 1: factor_total *= 0.80; ajustes.append("Mes Enero (0.80)")
    elif month == 12: factor_total *= 0.80; ajustes.append("Mes Diciembre (0.80)")
    cursor.execute("SELECT fecha FROM public.feriados WHERE fecha = %s", (fecha,))
    if cursor.fetchone():
        if dia_semana == 5: factor_total *= 0.50; ajustes.append("Feriado Sábado (0.50)")
        else: factor_total *= 0.30; ajustes.append("Feriado Laboral (0.30)")
    else:
        fecha_ant = fecha - timedelta(days=1)
        fecha_sig = fecha + timedelta(days=1)
        cursor.execute(
            "SELECT fecha FROM public.feriados WHERE fecha IS NOT NULL AND (fecha = %s OR fecha = %s)",
            (fecha_ant, fecha_sig)
        )
        feriados_cercanos = cursor.fetchall()
        def _fecha_val(r):
            return r[0] if isinstance(r, (list, tuple)) else r.get('fecha')
        pre = any(_fecha_val(f) == fecha_sig for f in feriados_cercanos)
        post = any(_fecha_val(f) == fecha_ant for f in feriados_cercanos)
        if post: factor_total *= 0.70; ajustes.append("Post-Feriado (0.70)")
        if pre: factor_total *= 0.70; ajustes.append("Pre-Feriado (0.70)")
    try:
        cursor.execute(
            "SELECT mm_caidos FROM control_metricas.t_casuisticas_lluvia WHERE fecha_evento=%s AND mm_caidos > 5 ORDER BY mm_caidos DESC LIMIT 1",
            (fecha,)
        )
        lluvia = cursor.fetchone()
        if lluvia:
            factor_total *= 0.50
            precipitacion = lluvia.get('mm_caidos', lluvia[0] if isinstance(lluvia, (list, tuple)) else 0)
            ajustes.append(f"Lluvia {precipitacion}mm (0.50)")
    except Exception:
        pass
    return round(factor_total, 2), ajustes


def get_tipo_dia_and_franjas(cursor, fecha: date) -> tuple:
    cursor.execute("SELECT fecha FROM public.feriados WHERE fecha = %s", (fecha,))
    es_feriado = cursor.fetchone() is not None
    dia_semana = fecha.weekday()
    if es_feriado or dia_semana == 6:
        id_tipo_dia, tipo_dia = 7, 'FERIADO' if es_feriado else 'NO LABORAL'
    elif dia_semana == 5:
        id_tipo_dia, tipo_dia = 6, 'SABADO'
    else:
        id_tipo_dia, tipo_dia = 5, 'LABORAL'
    cursor.execute("""
        SELECT id_franja, denominacion, hora_inicio, hora_fin
        FROM control_metricas.franjas_operativas
        WHERE id_tipo_dia = %s
          AND (inicio_vigencia IS NULL OR inicio_vigencia <= %s)
          AND (fin_vigencia IS NULL OR fin_vigencia >= %s)
        ORDER BY hora_inicio
    """, (id_tipo_dia, fecha, fecha))
    franjas = [dict(r) for r in cursor.fetchall()]
    return id_tipo_dia, tipo_dia, franjas


def get_cbd_for_hour(cod_catalogo: int, id_eot_vmt_hex: str, fecha, hora: int,
                     cbd_min_hora: int, cache_val: dict, cache_gps: dict) -> int:
    cbd_val = cache_val.get((cod_catalogo, fecha, hora), 0) or 0
    if cbd_val < cbd_min_hora and id_eot_vmt_hex:
        cbd_gps = cache_gps.get((id_eot_vmt_hex, fecha, hora), 0) or 0
        return max(cbd_val, cbd_gps)
    return cbd_val


def calcular_cbd_objetivo_hora(
    cod_catalogo: int, id_eot_vmt_hex: str, fecha: date, hora: int,
    cbd_min_hora: int, fechas_historicas: list, factor_ajuste: float,
    cache_val: dict, cache_gps: dict
) -> int:
    sum_hist = count = 0
    for fh in fechas_historicas:
        cbd_h = get_cbd_for_hour(
            cod_catalogo, id_eot_vmt_hex, fh, hora,
            cbd_min_hora, cache_val, cache_gps
        )
        sum_hist += cbd_h
        count += 1
    promedio_raw = sum_hist / count if count > 0 else 0
    b_dist_ajustado = max(promedio_raw * factor_ajuste, cbd_min_hora)
    cbd_objetivo = math.ceil(1.00 * b_dist_ajustado)
    return max(cbd_objetivo, cbd_min_hora)


def calcular_datos_cbd(cursor, fechas: list, eots: list, cache_val: dict, cache_gps: dict):
    """
    Calcula CBD objetivo para todas las empresas y fechas.
    Retorna: datos[cod_catalogo][fecha][hora] = cbd_objetivo,
             fechas_historicas_por_fecha, horas_posibles
    """
    horas_posibles = list(range(4, 23))
    datos = {}  # cod_catalogo -> fecha -> hora -> cbd_objetivo
    fechas_historicas_por_fecha = {}

    for eot in eots:
        cod_catalogo = eot['cod_catalogo']
        id_eot_vmt_hex = eot.get('id_eot_vmt_hex') or ''
        datos[cod_catalogo] = {}

        for fecha in fechas:
            id_tipo_dia, tipo_dia, franjas = get_tipo_dia_and_franjas(cursor, fecha)
            fechas_historicas = get_fechas_referencia(cursor, fecha)
            fechas_historicas_por_fecha[fecha] = fechas_historicas
            factor_ajuste, _ = get_factores_ajuste_acumulados(cursor, fecha)

            params_por_hora = {}
            for fr in franjas:
                h_ini = fr['hora_inicio'].hour
                h_fin = fr['hora_fin'].hour
                cursor.execute("""
                    SELECT cbd_minimo_hora FROM control_metricas.cbd_parametros_minimos
                    WHERE id_franja = %s AND (vigencia_desde IS NULL OR vigencia_desde <= %s)
                      AND (vigencia_hasta IS NULL OR vigencia_hasta >= %s)
                    LIMIT 1
                """, (fr['id_franja'], fecha, fecha))
                p = cursor.fetchone()
                cbd_min = p['cbd_minimo_hora'] if p else 0
                for h in range(h_ini, h_fin + 1):
                    params_por_hora[h] = max(params_por_hora.get(h, 0), cbd_min)

            datos[cod_catalogo][fecha] = {}
            for hora in horas_posibles:
                if hora in params_por_hora:
                    cbd_min_hora = params_por_hora[hora]
                    datos[cod_catalogo][fecha][hora] = calcular_cbd_objetivo_hora(
                        cod_catalogo, id_eot_vmt_hex, fecha, hora,
                        cbd_min_hora, fechas_historicas, factor_ajuste,
                        cache_val, cache_gps
                    )
                else:
                    datos[cod_catalogo][fecha][hora] = None

    return datos, fechas_historicas_por_fecha, horas_posibles


def _tabla_html_cbd(datos_eot: dict, fechas: list, horas_posibles: list, solo_valores=True) -> str:
    """Genera HTML de tabla CBD para un EOT o para todos."""
    rows = []
    rows.append("<tr><th>Fecha</th>" + "".join(f"<th>{h}:00</th>" for h in horas_posibles) + "</tr>")
    for fecha in fechas:
        row = f"<tr><td><strong>{fecha.strftime('%Y-%m-%d')}</strong></td>"
        for h in horas_posibles:
            val = datos_eot.get(fecha, {}).get(h)
            row += f"<td>{val if val is not None else '-'}</td>"
        row += "</tr>"
        rows.append(row)
    return f'<table border="1" cellpadding="5" cellspacing="0" style="border-collapse:collapse;font-size:12px;"><thead>{rows[0]}</thead><tbody>' + "".join(rows[1:]) + "</tbody></table>"


DIAS_SEMANA = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']


def _formatear_fechas_historicas(fechas_historicas_por_fecha: dict) -> str:
    lineas = []
    for fecha, fh_list in sorted(fechas_historicas_por_fecha.items()):
        fh_str = ", ".join(d.strftime("%d/%m/%Y") for d in sorted(fh_list))
        dia_nom = DIAS_SEMANA[fecha.weekday()]
        lineas.append(f"  • {fecha.strftime('%d/%m/%Y')} ({dia_nom}): {fh_str}")
    return "\n".join(lineas) if lineas else "  (No aplica)"


# =============================================================================
# ENVÍO DE EMAIL
# =============================================================================

def _get_smtp_config():
    """Obtiene configuración SMTP desde .env."""
    smtp_host = os.getenv("EMAIL_HOST") or os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.getenv("EMAIL_PORT") or os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("EMAIL_USER") or os.getenv("EMAIL_HOST_USER") or os.getenv("EMAIL_USERNAME", "")
    smtp_pass = os.getenv("EMAIL_PASSWORD") or os.getenv("EMAIL_HOST_PASSWORD", "")
    email_from = os.getenv("EMAIL_FROM", "")
    email_director = os.getenv("EMAIL_DIRECTOR") or os.getenv("EMAIL_TO", "")
    use_tls = os.getenv("EMAIL_USE_TLS", "true").lower() in ("true", "1", "t", "yes")
    return smtp_host, smtp_port, smtp_user, smtp_pass, email_from, email_director, use_tls


def enviar_email(destinatario: str, asunto: str, cuerpo_html: str) -> bool:
    """Envía un correo electrónico vía SMTP."""
    smtp_host, smtp_port, smtp_user, smtp_pass, email_from, _, use_tls = _get_smtp_config()
    if not all([smtp_host, smtp_user, smtp_pass, email_from]):
        print("✗ Error: Faltan variables de email en .env (EMAIL_HOST, EMAIL_USER, EMAIL_PASSWORD, EMAIL_FROM)")
        return False
    try:
        msg = MIMEMultipart("alternative")
        msg["From"] = email_from
        msg["To"] = destinatario
        msg["Subject"] = asunto
        msg.attach(MIMEText(cuerpo_html, "html"))
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=15)
        if use_tls:
            server.starttls()
        server.login(smtp_user, smtp_pass.strip())
        server.sendmail(email_from, [destinatario], msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print(f"✗ Error enviando correo a {destinatario}: {e}")
        return False


# =============================================================================
# GENERAR EXCEL
# =============================================================================

def generar_excel(ano: int = 2026, salida: str = "cbd_objetivo_feb.xlsx", datos_contexto=None) -> str:
    """Genera el Excel con una hoja por empresa."""
    try:
        import openpyxl  # type: ignore[import-untyped]
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # type: ignore[import-untyped]
    except ImportError:
        print("Error: Se requiere openpyxl. Ejecute: pip install openpyxl")
        sys.exit(1)

    fechas = [date(ano, 2, d) for d in (1, 2, 3, 4, 5)]
    cursor, conn = get_db_cursor()
    try:
        try:
            cursor.execute("""
                SELECT eot_id, cod_catalogo, eot_nombre, id_eot_vmt_hex, e_mail
                FROM public.eots
                WHERE cod_catalogo NOT IN (72)
                  AND permisionario IS TRUE
                ORDER BY cod_catalogo
            """)
            eots = [dict(r) for r in cursor.fetchall()]
        except Exception:
            cursor.execute("""
                SELECT eot_id, cod_catalogo, eot_nombre, id_eot_vmt_hex
                FROM public.eots
                WHERE cod_catalogo NOT IN (72)
                  AND permisionario IS TRUE
                ORDER BY cod_catalogo
            """)
            eots = [dict(r) for r in cursor.fetchall()]
            for eot in eots:
                eot['e_mail'] = None

        if not eots:
            print("No se encontraron EOTs en la base de datos.")
            return ""

        horas_posibles = list(range(4, 23))
        todas_fechas = list(fechas)
        for f in fechas:
            todas_fechas.extend(get_fechas_referencia(cursor, f))
        todas_fechas = list(set(todas_fechas))

        cod_catalogos = [e['cod_catalogo'] for e in eots]
        vmt_hexs = [e['id_eot_vmt_hex'] for e in eots if e.get('id_eot_vmt_hex')]

        cursor.execute("""
            SELECT id_eot_catalogo, fecha, hora, COUNT(DISTINCT idsam) as cbd
            FROM public.servicios_diarios
            WHERE id_eot_catalogo = ANY(%s) AND fecha = ANY(%s)
            GROUP BY id_eot_catalogo, fecha, hora
        """, (cod_catalogos, todas_fechas))
        cache_val = {(r['id_eot_catalogo'], r['fecha'], r['hora']): r['cbd'] for r in cursor.fetchall()}

        cache_gps = {}
        if vmt_hexs:
            try:
                cursor.execute("""
                    SELECT id_eot_vmt_hex, fecha, hora, COUNT(DISTINCT mean_id) as cbd
                    FROM control_metricas.cbd_detalle_buses
                    WHERE id_eot_vmt_hex = ANY(%s) AND fecha = ANY(%s)
                    GROUP BY id_eot_vmt_hex, fecha, hora
                """, (vmt_hexs, todas_fechas))
                cache_gps = {(r['id_eot_vmt_hex'], r['fecha'], r['hora']): r['cbd'] for r in cursor.fetchall()}
            except Exception:
                pass

        datos, fechas_historicas_por_fecha, _ = calcular_datos_cbd(cursor, fechas, eots, cache_val, cache_gps)

        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for eot in eots:
            cod_catalogo = eot['cod_catalogo']
            eot_nombre = eot['eot_nombre']
            sheet_name = (eot_nombre or f"EOT_{cod_catalogo}")[:31]
            for c in r':\/*?[]':
                sheet_name = sheet_name.replace(c, '_')

            ws = wb.create_sheet(title=sheet_name)
            ws.cell(row=1, column=1, value="Fecha")
            for col, h in enumerate(horas_posibles, start=2):
                ws.cell(row=1, column=col, value=f"{h}:00")
            for col in range(1, len(horas_posibles) + 2):
                c = ws.cell(row=1, column=col)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal='center', wrap_text=True)
                c.border = thin_border

            datos_eot = datos.get(cod_catalogo, {})
            for row_idx, fecha in enumerate(fechas, start=2):
                ws.cell(row=row_idx, column=1, value=fecha.strftime("%Y-%m-%d")).border = thin_border
                for col_idx, hora in enumerate(horas_posibles, start=2):
                    val = datos_eot.get(fecha, {}).get(hora)
                    ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "-")
                    ws.cell(row=row_idx, column=col_idx).border = thin_border
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')

            ws.column_dimensions['A'].width = 12
            for col in range(2, len(horas_posibles) + 2):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 6

        ruta_salida = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", salida))
        wb.save(ruta_salida)
        print(f"Excel generado correctamente: {ruta_salida}")
        print(f"  - {len(eots)} hojas | Fechas: {fechas[0]} al {fechas[-1]} | Columnas: 4:00 a 22:00")
        return ruta_salida
    finally:
        cursor.close()
        conn.close()


# =============================================================================
# VERIFICACIÓN (DIRECTOR) Y NOTIFICACIÓN (EMPRESAS)
# =============================================================================

def ejecutar_verificacion(ano: int = 2026):
    """Envía correo al director con CBD objetivo de todas las empresas."""
    fechas = [date(ano, 2, d) for d in (1, 2, 3, 4, 5)]
    email_director = os.getenv("EMAIL_DIRECTOR") or os.getenv("EMAIL_TO", "")
    if not email_director:
        print("✗ Error: Configure EMAIL_DIRECTOR o EMAIL_TO en .env para --verificacion")
        return False

    cursor, conn = get_db_cursor()
    try:
        try:
            cursor.execute("""
                SELECT eot_id, cod_catalogo, eot_nombre, id_eot_vmt_hex, e_mail
                FROM public.eots WHERE cod_catalogo NOT IN (72) AND permisionario IS TRUE
                ORDER BY cod_catalogo
            """)
        except Exception:
            cursor.execute("""
                SELECT eot_id, cod_catalogo, eot_nombre, id_eot_vmt_hex
                FROM public.eots WHERE cod_catalogo NOT IN (72) AND permisionario IS TRUE
                ORDER BY cod_catalogo
            """)
        eots = [dict(r) for r in cursor.fetchall()]
        for eot in eots:
            if 'e_mail' not in eot:
                eot['e_mail'] = None

        if not eots:
            print("No se encontraron EOTs.")
            return False

        horas_posibles = list(range(4, 23))
        todas_fechas = list(fechas)
        for f in fechas:
            todas_fechas.extend(get_fechas_referencia(cursor, f))
        todas_fechas = list(set(todas_fechas))

        cod_catalogos = [e['cod_catalogo'] for e in eots]
        vmt_hexs = [e['id_eot_vmt_hex'] for e in eots if e.get('id_eot_vmt_hex')]

        cursor.execute("""SELECT id_eot_catalogo, fecha, hora, COUNT(DISTINCT idsam) as cbd
            FROM public.servicios_diarios WHERE id_eot_catalogo = ANY(%s) AND fecha = ANY(%s)
            GROUP BY id_eot_catalogo, fecha, hora""", (cod_catalogos, todas_fechas))
        cache_val = {(r['id_eot_catalogo'], r['fecha'], r['hora']): r['cbd'] for r in cursor.fetchall()}

        cache_gps = {}
        if vmt_hexs:
            try:
                cursor.execute("""SELECT id_eot_vmt_hex, fecha, hora, COUNT(DISTINCT mean_id) as cbd
                    FROM control_metricas.cbd_detalle_buses
                    WHERE id_eot_vmt_hex = ANY(%s) AND fecha = ANY(%s)
                    GROUP BY id_eot_vmt_hex, fecha, hora""", (vmt_hexs, todas_fechas))
                cache_gps = {(r['id_eot_vmt_hex'], r['fecha'], r['hora']): r['cbd'] for r in cursor.fetchall()}
            except Exception:
                pass

        datos, fechas_historicas_por_fecha, _ = calcular_datos_cbd(cursor, fechas, eots, cache_val, cache_gps)

        fechas_historicas_texto = _formatear_fechas_historicas(fechas_historicas_por_fecha)

        tablas = []
        for eot in eots:
            cod = eot['cod_catalogo']
            nombre = eot['eot_nombre'] or f"EOT_{cod}"
            tablas.append(f"<h4>{nombre}</h4>" + _tabla_html_cbd(datos.get(cod, {}), fechas, list(range(4, 23))))

        borrador_empresa_html = TEXTO_BASE_EMPRESA.format(
            eot_nombre="[NOMBRE_EMPRESA]",
            ano=ano,
            fechas_historicas_texto=fechas_historicas_texto.replace("\n", "<br>"),
            cbd_tabla_html='<p style="background:#f0f0f0;padding:10px;font-style:italic;">[En el correo a cada empresa se inserta aquí la tabla con el CBD Objetivo correspondiente a esa empresa]</p>'
        ).replace("\n", "<br>")

        cuerpo = f"""
        <html><body style="font-family:Arial,sans-serif;">
        <h2>CBD Objetivo - Verificación para Director</h2>
        <p>Período: 1 al 5 de febrero de {ano}</p>
        """ + TEXTO_VERIFICACION_DIRECTOR.format(
            ano=ano,
            fechas_historicas_texto=fechas_historicas_texto.replace("\n", "<br>"),
            cbd_tabla_todas_html="<br>".join(tablas)
        ) + """
        <hr style="margin:30px 0;border:none;border-top:1px solid #ccc;">
        <h3>Borrador de Texto para empresa</h3>
        <p style="font-size:12px;color:#666;">El siguiente es el texto que recibirá cada empresa en su correo de notificación (con su nombre y tabla CBD específica):</p>
        <div style="background:#fafafa;border:1px solid #ddd;padding:20px;margin:15px 0;font-size:13px;line-height:1.5;">
        """ + borrador_empresa_html + """
        </div>
        </body></html>"""

        if enviar_email(
            email_director,
            f"[DMT] CBD Objetivo - Verificación empresas (1-5 feb {ano})",
            cuerpo
        ):
            print(f"✓ Correo de verificación enviado al director ({email_director})")
            return True
        return False
    finally:
        cursor.close()
        conn.close()


def ejecutar_notificacion(ano: int = 2026):
    """Envía correo a cada empresa permisionaria con su CBD objetivo."""
    fechas = [date(ano, 2, d) for d in (1, 2, 3, 4, 5)]
    cursor, conn = get_db_cursor()

    try:
        try:
            cursor.execute("""
                SELECT eot_id, cod_catalogo, eot_nombre, id_eot_vmt_hex, e_mail
                FROM public.eots WHERE cod_catalogo NOT IN (72) AND permisionario IS TRUE
                ORDER BY cod_catalogo
            """)
        except Exception:
            cursor.execute("""
                SELECT eot_id, cod_catalogo, eot_nombre, id_eot_vmt_hex
                FROM public.eots WHERE cod_catalogo NOT IN (72) AND permisionario IS TRUE
                ORDER BY cod_catalogo
            """)
        eots = [dict(r) for r in cursor.fetchall()]
        for eot in eots:
            if 'e_mail' not in eot:
                eot['e_mail'] = None

        eots_con_email = [e for e in eots if e.get('e_mail')]
        if not eots_con_email:
            print("✗ No hay EOTs con e_mail configurado en la base de datos.")
            return False

        horas_posibles = list(range(4, 23))
        todas_fechas = list(fechas)
        for f in fechas:
            todas_fechas.extend(get_fechas_referencia(cursor, f))
        todas_fechas = list(set(todas_fechas))

        cod_catalogos = [e['cod_catalogo'] for e in eots]
        vmt_hexs = [e['id_eot_vmt_hex'] for e in eots if e.get('id_eot_vmt_hex')]

        cursor.execute("""SELECT id_eot_catalogo, fecha, hora, COUNT(DISTINCT idsam) as cbd
            FROM public.servicios_diarios WHERE id_eot_catalogo = ANY(%s) AND fecha = ANY(%s)
            GROUP BY id_eot_catalogo, fecha, hora""", (cod_catalogos, todas_fechas))
        cache_val = {(r['id_eot_catalogo'], r['fecha'], r['hora']): r['cbd'] for r in cursor.fetchall()}

        cache_gps = {}
        if vmt_hexs:
            try:
                cursor.execute("""SELECT id_eot_vmt_hex, fecha, hora, COUNT(DISTINCT mean_id) as cbd
                    FROM control_metricas.cbd_detalle_buses
                    WHERE id_eot_vmt_hex = ANY(%s) AND fecha = ANY(%s)
                    GROUP BY id_eot_vmt_hex, fecha, hora""", (vmt_hexs, todas_fechas))
                cache_gps = {(r['id_eot_vmt_hex'], r['fecha'], r['hora']): r['cbd'] for r in cursor.fetchall()}
            except Exception:
                pass

        datos, fechas_historicas_por_fecha, _ = calcular_datos_cbd(cursor, fechas, eots, cache_val, cache_gps)
        fechas_historicas_texto = _formatear_fechas_historicas(fechas_historicas_por_fecha)
        fechas_historicas_texto_legible = "\n".join(
            f"  • {fecha.strftime('%d/%m/%Y')} ({DIAS_SEMANA[fecha.weekday()]}): " +
            ", ".join(d.strftime("%d/%m/%Y") for d in sorted(fechas_historicas_por_fecha[fecha]))
            for fecha in sorted(fechas_historicas_por_fecha)
        )

        ok, fallidos = 0, []
        for eot in eots_con_email:
            cod = eot['cod_catalogo']
            nombre = eot['eot_nombre'] or f"EOT_{cod}"
            email = eot['e_mail']
            tabla_html = _tabla_html_cbd(datos.get(cod, {}), fechas, horas_posibles)

            cuerpo = f"""
            <html><body style="font-family:Arial,sans-serif;font-size:14px;">
            """ + TEXTO_BASE_EMPRESA.format(
                eot_nombre=nombre,
                ano=ano,
                fechas_historicas_texto=fechas_historicas_texto_legible.replace("\n", "<br>"),
                cbd_tabla_html=tabla_html
            ) + "</body></html>"

            if enviar_email(
                email,
                f"[DMT] CBD Objetivo - {nombre} (1-5 feb {ano})",
                cuerpo
            ):
                ok += 1
                print(f"  ✓ {nombre} -> {email}")
            else:
                fallidos.append(nombre)

        print(f"\n✓ Notificaciones enviadas: {ok}/{len(eots_con_email)}")
        if fallidos:
            print(f"  Fallidos: {', '.join(fallidos)}")
        return ok == len(eots_con_email)
    finally:
        cursor.close()
        conn.close()


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generar Excel y/o enviar correos con CBD objetivo por empresa (IFO 100%)"
    )
    parser.add_argument("--ano", type=int, default=2026, help="Año (default: 2026)")
    parser.add_argument("--salida", type=str, default="cbd_objetivo_feb.xlsx", help="Nombre del archivo Excel")
    parser.add_argument("--verificacion", action="store_true",
                        help="Envía correo al director con CBD objetivo de todas las empresas")
    parser.add_argument("--notificacion", action="store_true",
                        help="Envía correo a cada empresa con su CBD objetivo")
    args = parser.parse_args()

    if args.verificacion:
        ejecutar_verificacion(ano=args.ano)
    elif args.notificacion:
        ejecutar_notificacion(ano=args.ano)
    else:
        generar_excel(ano=args.ano, salida=args.salida)


if __name__ == "__main__":
    main()
