import sys
import os
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Asegurar que importamos del backend correctamente
sys.path.append(r'c:\Users\rafael\Documents\Desarrollos\CID\ifo\ifo\cbd_monitor\backend')
from database.connection import get_db_connection

def generate_report():
    print("Conectando a la base de datos...")
    db = next(get_db_connection())
    cursor = db.get_cursor()

    start_date = date(2025, 11, 24)
    end_date = date(2026, 5, 31)

    print("Obteniendo franjas operativas...")
    # Obtenemos franjas operativas activas
    cursor.execute("SELECT id_franja, denominacion, hora_inicio, hora_fin, id_tipo_dia FROM control_metricas.franjas_operativas WHERE activo = TRUE ORDER BY hora_inicio")
    franjas_db = cursor.fetchall()

    # Setup HTML
    html = "<html><head><meta charset='UTF-8'><title>Reporte Gonzalez Quiñonez</title>"
    html += "<style>table { border-collapse: collapse; width: 100%; font-family: sans-serif; font-size: 14px; } th, td { border: 1px solid #ddd; padding: 8px; text-align: center; } th { background-color: #f2f2f2; } .fecha-row { background-color: #e6f7ff; font-weight: bold; }</style>"
    html += "</head><body><h2>Reporte CBD por Franja - Gonzalez Quiñonez (24/11/2025 al 31/05/2026)</h2>"
    html += "<table>"

    # Setup Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte_GQ"
    
    # Estilos Excel
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    fecha_fill = PatternFill("solid", fgColor="E6F7FF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(["Reporte CBD por Franja - Gonzalez Quiñonez"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])

    current_date = start_date
    total_dias = (end_date - start_date).days + 1
    dia_actual = 1

    row_excel = 3

    while current_date <= end_date:
        if dia_actual % 10 == 0:
            print(f"Procesando día {dia_actual} de {total_dias} ({current_date})...")
            
        dia_semana = current_date.weekday()
        if dia_semana <= 4:
            id_tipo_dia = 5
        elif dia_semana == 5:
            id_tipo_dia = 6
        else:
            id_tipo_dia = 7

        franjas_dia = [f for f in franjas_db if f['id_tipo_dia'] == id_tipo_dia]
        if not franjas_dia:
            current_date += timedelta(days=1)
            dia_actual += 1
            continue

        # HTML: Fila Fecha
        html += f"<tr class='fecha-row'><td colspan='{len(franjas_dia) + 2}' style='text-align: left;'>Fecha: {current_date.strftime('%d/%m/%Y')}</td></tr>"
        
        # Excel: Fila Fecha
        fecha_text = f"Fecha: {current_date.strftime('%d/%m/%Y')}"
        ws.append([fecha_text])
        ws.merge_cells(start_row=row_excel, start_column=1, end_row=row_excel, end_column=len(franjas_dia)+2)
        cell = ws.cell(row=row_excel, column=1)
        cell.font = header_font
        cell.fill = fecha_fill
        cell.alignment = left_align
        for col in range(1, len(franjas_dia) + 3):
            ws.cell(row=row_excel, column=col).border = thin_border
        row_excel += 1

        # Headers franjas
        # HTML
        html += "<tr><th>Métrica</th>"
        for f in franjas_dia:
            html += f"<th>{f['denominacion']}<br><small>{f['hora_inicio'].strftime('%H:%M')} - {f['hora_fin'].strftime('%H:%M')}</small></th>"
        html += "<th>Total</th></tr>"
        
        # Excel
        header_row = ["Métrica"]
        for f in franjas_dia:
            header_row.append(f"{f['denominacion']}\n{f['hora_inicio'].strftime('%H:%M')} - {f['hora_fin'].strftime('%H:%M')}")
        header_row.append("Total")
        ws.append(header_row)
        for col in range(1, len(header_row) + 1):
            c = ws.cell(row=row_excel, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border
        row_excel += 1

        # 1) Validaciones
        html += "<tr><td>Validaciones</td>"
        row_val = ["Validaciones"]
        val_counts = {}
        cursor.execute('''
            SELECT 
                fo.id_franja,
                COUNT(DISTINCT sd.idsam) as cantidad
            FROM public.servicios_diarios sd
            INNER JOIN control_metricas.franjas_operativas fo 
                ON (sd.hora * INTERVAL '1 hour')::time >= fo.hora_inicio 
                AND (sd.hora * INTERVAL '1 hour')::time < fo.hora_fin
            WHERE sd.id_eot_catalogo = 11 AND sd.fecha = %s AND fo.id_tipo_dia = %s AND fo.activo = TRUE
            GROUP BY fo.id_franja
        ''', (current_date, id_tipo_dia))
        for row in cursor.fetchall():
            val_counts[row['id_franja']] = row['cantidad']
        
        cursor.execute('''
            SELECT COUNT(DISTINCT idsam) as total
            FROM public.servicios_diarios
            WHERE id_eot_catalogo = 11 AND fecha = %s
        ''', (current_date,))
        total_val = cursor.fetchone()['total']

        for f in franjas_dia:
            v = val_counts.get(f['id_franja'], 0)
            html += f"<td>{v}</td>"
            row_val.append(v)
            
        html += f"<td><b>{total_val}</b></td></tr>"
        row_val.append(total_val)
        
        ws.append(row_val)
        for col in range(1, len(row_val) + 1):
            c = ws.cell(row=row_excel, column=col)
            c.alignment = center_align
            c.border = thin_border
            if col == len(row_val):
                c.font = Font(bold=True)
        row_excel += 1

        # 2) GPS
        html += "<tr><td>GPS</td>"
        row_gps = ["GPS"]
        gps_counts = {}
        cursor.execute('''
            SELECT 
                fo.id_franja,
                COUNT(DISTINCT cdb.mean_id) as cantidad
            FROM control_metricas.cbd_detalle_buses cdb
            INNER JOIN control_metricas.franjas_operativas fo 
                ON (cdb.hora * INTERVAL '1 hour')::time >= fo.hora_inicio 
                AND (cdb.hora * INTERVAL '1 hour')::time < fo.hora_fin
            WHERE cdb.id_eot_vmt_hex = '000B' AND cdb.fecha = %s AND fo.id_tipo_dia = %s AND fo.activo = TRUE
            GROUP BY fo.id_franja
        ''', (current_date, id_tipo_dia))
        for row in cursor.fetchall():
            gps_counts[row['id_franja']] = row['cantidad']
        
        cursor.execute('''
            SELECT COUNT(DISTINCT mean_id) as total
            FROM control_metricas.cbd_detalle_buses
            WHERE id_eot_vmt_hex = '000B' AND fecha = %s
        ''', (current_date,))
        total_gps = cursor.fetchone()['total']

        for f in franjas_dia:
            v = gps_counts.get(f['id_franja'], 0)
            html += f"<td>{v}</td>"
            row_gps.append(v)
            
        html += f"<td><b>{total_gps}</b></td></tr>"
        row_gps.append(total_gps)
        
        ws.append(row_gps)
        for col in range(1, len(row_gps) + 1):
            c = ws.cell(row=row_excel, column=col)
            c.alignment = center_align
            c.border = thin_border
            if col == len(row_gps):
                c.font = Font(bold=True)
        row_excel += 1

        # 3) Servicios (public.nservicios)
        html += "<tr><td>Servicios (GQ)</td>"
        row_serv = ["Servicios (GQ)"]
        
        cursor.execute('''
            SELECT mean_id, inicio_serv, fin_serv
            FROM public.nservicios
            WHERE servicio = 'GQ' AND fecha_operativa = %s
        ''', (current_date,))
        servicios = cursor.fetchall()
        
        buses_por_franja = {f['id_franja']: set() for f in franjas_dia}
        buses_totales = set()
        
        for s in servicios:
            if not s['inicio_serv'] or not s['fin_serv']:
                continue
            
            t_inicio = s['inicio_serv'].time()
            t_fin = s['fin_serv'].time()
            buses_totales.add(s['mean_id'])
            
            for f in franjas_dia:
                if t_inicio <= f['hora_fin'] and t_fin >= f['hora_inicio']:
                    buses_por_franja[f['id_franja']].add(s['mean_id'])
                    
        for f in franjas_dia:
            v = len(buses_por_franja[f['id_franja']])
            html += f"<td>{v}</td>"
            row_serv.append(v)
            
        total_serv = len(buses_totales)
        html += f"<td><b>{total_serv}</b></td></tr>"
        row_serv.append(total_serv)
        
        ws.append(row_serv)
        for col in range(1, len(row_serv) + 1):
            c = ws.cell(row=row_excel, column=col)
            c.alignment = center_align
            c.border = thin_border
            if col == len(row_serv):
                c.font = Font(bold=True)
        row_excel += 1

        current_date += timedelta(days=1)
        dia_actual += 1
        
        # Add a blank row in Excel for spacing
        ws.append([])
        row_excel += 1

    html += "</table></body></html>"

    output_path_html = r'c:\Users\rafael\Documents\Desarrollos\CID\ifo\ifo\cbd_monitor\reporte_gq.html'
    output_path_excel = r'c:\Users\rafael\Documents\Desarrollos\CID\ifo\ifo\cbd_monitor\reporte_gq.xlsx'
    
    with open(output_path_html, "w", encoding="utf-8") as file:
        file.write(html)
        
    # Auto-adjust column widths in Excel
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
        
    wb.save(output_path_excel)
    
    cursor.close()
    print(f"Reporte generado exitosamente:")
    print(f" HTML: {output_path_html}")
    print(f" Excel: {output_path_excel}")

if __name__ == '__main__':
    generate_report()
